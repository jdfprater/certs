import openpyxl, time, pprint, json, re
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

start_time = time.time()
print('Starting...')

GYMS =              ['TX-AUSTIN ANDERSON ARBOR',
                      'TX-AUSTIN CEDAR PARK',
                      'TX-AUSTIN CYPRESS CREEK',
                      'TX-AUSTIN HESTERS CROSSING',
                      'TX-AUSTIN NORTH ROUND ROCK',
                      'TX-AUSTIN TECHRIDGE',
                      'TX-GEORGETOWN',
                      'TX-PFLUGERVILLE']

PT_Data = {}

print('Opening Payroll Report.')
twb = openpyxl.load_workbook('PT Training Payroll Report.xlsx')
tws = twb['PT_Payroll_Summary']
print('That thing was massive and thats what she said.')


def noMiddle(name):
    middleRegex = re.compile(r'\D*,\s\D*\s\D{1}')
    mo = middleRegex.search(name)
    if mo:
        name = name[:-2]
    return name.upper()

def lastFirst(first, last):
    lastFirst = str(last).upper() + ', ' + str(first).upper()
    return lastFirst

trainer = column_index_from_string('E')
bonus_hours = column_index_from_string('J')
club = column_index_from_string('G')

for row in tws.rows:
    pt = noMiddle(str(row[trainer-1].value))
    hours = row[bonus_hours-1].value
    gym = row[club-1].value
    if pt:
        if gym in GYMS and pt in PT_Data:
            PT_Data[pt]['bonus hours'] += hours
        elif gym in GYMS and pt not in PT_Data:
            PT_Data.update( {pt : {'bonus hours' : hours, 'cpt' : [], 'sales' : 0, 'degree' : '', 'gym' : gym}})
        else:
            continue
    else:
        continue

cwb = openpyxl.load_workbook('certs.xlsx')
cws = cwb.active

role = column_index_from_string('D')
first = column_index_from_string('E')
last = column_index_from_string('F')
agency = column_index_from_string('G')
certification = column_index_from_string('H')
fitness = ['Trainer', 'Assistant Fitness Manager', 'Fitness Manager']

for row in cws.rows:
    name = lastFirst(str(row[first-1].value), str(row[last-1].value))
    trainer = str(row[role-1].value)
    company = str(row[agency-1].value)
    cert = str(row[certification-1].value)
    if any(x in (trainer) for x in fitness):
        if 'Trainer' or 'CPT' in cert:
            if name in PT_Data:
                PT_Data[name]['cpt'].append(company)
            else:
                PT_Data.update( {name : {'bonus hours' : hours, 'cpt' : [], 'sales' : 0, 'degree' : '', 'gym' : gym, 'start' : ''}})
                PT_Data[name]['cpt'].append(company)
        elif 'Bachelor' in cert:
            if name in PT_Data:
                PT_Data[name]['degree'] = cert
            else:
                PT_Data.update( {name : {'bonus hours' : hours, 'cpt' : [], 'sales' : 0, 'degree' : cert, 'gym' : gym}})
                PT_Data[name]['degree'] = cert
        else:
            continue
    else:
        continue

headers = ['pt', 'hours', 'cpt', 'sales', 'degree', 'gym']
ptcerts = Workbook()
ptsheet = ptcerts.active

for i in range(len(headers)):
    ptsheet.cell(row=1, column=i+1).font = Font(bold=True)
    ptsheet.cell(row=1, column=i+1).value = headers[i]

row = 2
for pt, data in PT_Data.items():
    ptsheet.cell(row=row, column=1, value=pt)
    column = 2
    for k, v in data.items():
        ptsheet.cell(row=row, column=column, value=str(v))
        column += 1
    row += 1

ptcerts.save('pt_certs.xlsx')
ptcerts.close()

print("--- %s seconds ---" % (time.time() - start_time))

